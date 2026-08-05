# -*- coding: utf-8 -*-
"""Write menu-products.xlsx from menu.json - the sheet the menu is edited in.

menu.json is the source; this file is a view of it. The point of the view is
that the person who knows the prices does not have to open a JSON file to fix
one, so it carries the same strings the page carries and marks the cells that
are still waiting on real content.

Regenerate after any change to menu.json:  py menu_xlsx.py

Nothing reads this file back yet. Edits made in it are re-typed into menu.json
by hand, which is why every column that is not editable says so in the guide
sheet rather than being silently ignored.
"""
import io
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "menu.json")
OUT = os.path.join(HERE, "menu-products.xlsx")

# Placeholders left in menu.json where real copy has not been written yet.
# They ship to the page as-is, so the sheet has to show them as unfinished.
EMPTY_DESC = "[توضیح]"
EMPTY_INTRO = "[توضیح دسته]"
EMPTY_PRICE = "[قیمت]"

MAROON = "471019"   # brand ink, header band
CREAM = "F3F1EC"    # brand paper, header text
YELLOW = "FFFFCC"   # editable
ORANGE = "FFF3CD"   # editable and still empty

HEAD_FILL = PatternFill("solid", fgColor=MAROON)
HEAD_FONT = Font(name="Arial", size=11, bold=True, color=CREAM)
HEAD_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Arial", size=10)
HAIR = Side(style="thin")
BODY_BORDER = Border(top=HAIR, bottom=HAIR)

PRODUCTS = [
    # header,          width, align,    wrap,  editable
    ("ردیف",            6,  "center", False, False),
    ("دسته",           18,  "right",  False, False),
    ("نام محصول",      26,  "right",  False, True),
    ("کد سپیدز",       14,  "center", False, False),
    ("قیمت",           10,  "center", False, True),
    ("قیمت (عددی)",    12,  "center", False, False),
    ("گزینه‌ها",        30,  "right",  True,  True),
    ("توضیح",          62,  "right",  True,  True),
    ("وضعیت عکس",      11,  "center", False, False),
    ("فایل عکس",       26,  "right",  False, False),
    ("شناسه اسلات",    22,  "right",  False, False),
]
CATEGORIES = [
    ("ردیف",            6,  "center", False, False),
    ("نام دسته",       22,  "right",  False, False),
    ("توضیح دسته",     64,  "right",  True,  True),
    ("تعداد محصول",    13,  "center", False, False),
    ("شناسه دسته",     20,  "right",  False, False),
]

ONES = ["صفر", "یک", "دو", "سه", "چهار", "پنج", "شش", "هفت", "هشت", "نه"]
TEENS = ["ده", "یازده", "دوازده", "سیزده", "چهارده", "پانزده",
         "شانزده", "هفده", "هجده", "نوزده"]
TENS = ["", "", "بیست", "سی", "چهل", "پنجاه", "شصت", "هفتاد", "هشتاد", "نود"]


def fa_word(n):
    """Spell a small count, so the guide reads as a sentence and not a table."""
    if n < 10:
        return ONES[n]
    if n < 20:
        return TEENS[n - 10]
    if n >= 100:
        return str(n)
    tens, ones = divmod(n, 10)
    return TENS[tens] if not ones else TENS[tens] + "‌و" + ONES[ones]


def _items(data):
    """Every item, paired with the category it sits in."""
    return [(c, i) for c in data["categories"] for i in c["items"]]


def _codes(item):
    """Sepidz codes for one row, in the same order as its options."""
    if "codes" in item:
        return item["codes"]
    if "options" in item:
        return [o["code"] for o in item["options"] if "code" in o]
    return [item["code"]] if "code" in item else []


def _sheet(wb, title, columns, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.sheet_view.rightToLeft = True
    for n, (head, width, _, _, _) in enumerate(columns, 1):
        cell = ws.cell(1, n, head)
        cell.fill, cell.font, cell.alignment = HEAD_FILL, HEAD_FONT, HEAD_ALIGN
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    return ws


def _write(ws, row, columns, values, unfilled=()):
    """One body row, styled by column and marked where content is missing."""
    for n, ((_, _, align, wrap, editable), value) in enumerate(zip(columns, values), 1):
        cell = ws.cell(row, n, value)
        cell.font = BODY_FONT
        cell.border = BODY_BORDER
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        if n in unfilled:
            cell.fill = PatternFill("solid", fgColor=ORANGE)
        elif editable:
            cell.fill = PatternFill("solid", fgColor=YELLOW)


def build(data):
    wb = Workbook()

    # --- guide ------------------------------------------------------------
    ws = wb.active
    ws.title = "راهنما"
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 92

    pairs = _items(data)
    no_intro = sum(1 for c in data["categories"] if c["intro"] == EMPTY_INTRO)
    no_desc = sum(1 for _, i in pairs if i.get("desc") == EMPTY_DESC)
    no_price = sum(1 for _, i in pairs if i.get("price") == EMPTY_PRICE)
    no_photo = sum(1 for _, i in pairs if "slotId" in i and not i.get("photo"))

    opted = next(i for _, i in pairs if "options" in i)
    sample_cat, sample = next((c, i) for c, i in pairs if "slotId" in i
                              and i.get("photo") and i.get("price")
                              and i["desc"] != EMPTY_DESC)

    guide = [
        ("راهنمای این فایل", None),
        (None, None),
        ("منبع", "از menu.json ساخته شده؛ همان فایلی که صفحه منوی سایت از روی آن ساخته می‌شود."),
        ("خانه‌های زرد کم‌رنگ", "قابل ویرایش‌اند: نام محصول، قیمت، گزینه‌ها، توضیح، و توضیح دسته."),
        ("خانه‌های نارنجی", "هنوز پر نشده‌اند و همین حالا روی سایت دیده می‌شوند. در ستون وضعیت عکس یعنی"),
        (None, "عکس واقعی ندارد و عکس عمومی نشان داده می‌شود؛ در بقیه ستون‌ها یعنی متن داخل کروشه"),
        (None, "هنوز سر جایش مانده. مجموع: %s توضیح دسته، %s توضیح محصول، %s قیمت، و %s عکس."
               % (fa_word(no_intro), fa_word(no_desc), fa_word(no_price), fa_word(no_photo))),
        ("ستون قیمت", "با رقم فارسی و دقیقاً به همان شکلی که در سایت می‌آید. همین رشته به سایت برمی‌گردد،"),
        (None, "پس اگر با رقم انگلیسی بنویسید روی صفحه هم انگلیسی درمی‌آید."),
        ("ستون قیمت (عددی)", "فقط برای مرتب‌سازی و فیلتر است و به سایت برنمی‌گردد."),
        ("ستون گزینه‌ها", "برای محصولاتی که چند قیمت دارند، به شکل «برچسب: قیمت» و جداشده با علامت خط عمودی،"),
        (None, "مثل %s. این محصولات ستون قیمت‌شان خالی است." % _options(opted)),
        ("ستون کد سپیدز", "کد همین محصول در سپیدز. محصول چندقیمتی به ازای هر گزینه یک کد دارد، به همان"),
        (None, "ترتیب ستون گزینه‌ها. خالی یعنی در سپیدز چنین کدی پیدا نشد. فقط برای خواندن است."),
        ("شناسه اسلات", "کلید ثابت هر محصول برای وصل شدن به عکسش. دست نزنید."),
        (None, None),
        ("نمونه یک سطر پرشده", None),
        ("دسته", sample_cat["title"]),
        ("نام محصول", sample["name"]),
        ("کد سپیدز", " | ".join(_codes(sample))),
        ("قیمت", sample["price"]),
        ("توضیح", sample["desc"]),
        ("وضعیت عکس", "دارد"),
        ("شناسه اسلات", sample["slotId"]),
    ]
    for row, (key, text) in enumerate(guide, 1):
        a = ws.cell(row, 1, key)
        a.font = Font(name="Arial", size=14, bold=True, color=MAROON) if row == 1 \
            else Font(name="Arial", size=10, bold=True)
        a.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        b = ws.cell(row, 2, text)
        b.font = BODY_FONT
        b.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    # --- products ---------------------------------------------------------
    ws = _sheet(wb, "محصولات", PRODUCTS)
    for row, (cat, item) in enumerate(pairs, 2):
        options = _options(item)
        price = item.get("price", "")
        unfilled = []
        if price == EMPTY_PRICE:
            unfilled.append(5)
        if item.get("desc") == EMPTY_DESC:
            unfilled.append(8)
        if "slotId" in item and not item.get("photo"):
            unfilled.append(9)
        _write(ws, row, PRODUCTS, [
            row - 1,
            cat["title"],
            item["name"],
            " | ".join(_codes(item)),
            "" if options else price,
            "" if options else _digits(price),
            options,
            item.get("desc", ""),
            ("دارد" if item.get("photo") else "ندارد") if "slotId" in item else "—",
            item.get("photo", ""),
            item.get("slotId", ""),
        ], unfilled)
    last = len(pairs) + 1

    # --- categories -------------------------------------------------------
    ws = _sheet(wb, "دسته‌ها", CATEGORIES)
    for row, cat in enumerate(data["categories"], 2):
        _write(ws, row, CATEGORIES, [
            row - 1,
            cat["title"],
            cat["intro"],
            "=COUNTIF(محصولات!$B$2:$B$%d,B%d)" % (last, row),
            cat["id"],
        ], [3] if cat["intro"] == EMPTY_INTRO else [])
    total = len(data["categories"]) + 2
    _write(ws, total, CATEGORIES,
           ["", "جمع", "", "=SUM(D2:D%d)" % (total - 1), ""])

    return wb, len(pairs)


def _options(item):
    """The «label: price» string the sheet shows for a multi-price product."""
    return " | ".join("%s: %s" % (o["label"], o["price"]) for o in item.get("options", []))


def _digits(price):
    """The price again as a number, for sorting. Persian digits do not sort."""
    plain = price.translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789"))
    return int(plain) if plain.isdigit() else ""


def main():
    data = json.loads(io.open(SRC, encoding="utf-8").read())
    wb, count = build(data)
    wb.save(OUT)
    print("%s -> %s (%d products, %d categories)"
          % (os.path.basename(SRC), os.path.basename(OUT), count, len(data["categories"])))


if __name__ == "__main__":
    main()
